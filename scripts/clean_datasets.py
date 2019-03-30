import os
import argparse
from openpyxl import Workbook
from shutil import copyfile, rmtree


def modify_file(filename, worksheet=None, index=0, base_dir=None):
    print(f"File to clean: {filename}")
    attributes = []
    data = []
    attributes_used = []
    relation_name = ""
    with open(filename) as file:
        relation_name = file.readline()
        data_processing = False
        i = 0
        for line in file:
            if not data_processing:
                line_upper = line.upper()
                if line_upper.startswith("@ATTRIBUTE"):
                    attribute = line.split(' ')
                    name = attribute[1].rstrip()
                    att_type = ""
                    if len(attribute) > 2:
                        att_type = attribute[2].rstrip()
                    attributes.append((name, att_type, line, i))
                    i += 1
                    continue
                if line_upper.startswith("@INPUTS"):
                    inputs = [x.lstrip() for x in line.split(" ", 1)[1].split(',')]
                    attributes_used = [x.rstrip() for x in inputs]
                    continue
                if line_upper.startswith("@OUTPUT"):
                    outputs = [x.lstrip() for x in line.split(" ", 1)[1].split(',')]
                    attributes_used.extend([x.rstrip() for x in outputs])
                if line_upper.startswith("@DATA"):
                    data_processing = True
            else:
                line_data = [x.lstrip() for x in line.split(',')]
                data.append(line_data)

    permitted_attributes = [x for x in attributes if x[1] == "" and x[0].split('{')[0] in attributes_used]
    permitted_indexes = [x[3] for x in permitted_attributes]

    if len(permitted_attributes) > 1:
        new_filename = filename.replace(".dat", "_cleaned.arff")
        with open(new_filename, "w") as new_file:
            new_file.write(relation_name)
            for item in permitted_attributes:
                new_file.write(item[2].replace("{", " {"))
            new_file.write("@DATA\n")
            for datapoint in data:
                points = [x for i, x in enumerate(datapoint) if i in permitted_indexes]
                result = ",".join(points)
                new_file.write(result)
        print(f"Created file {new_filename}")
    simple_filename = filename.split('/')[-1].split('.')[0]
    if worksheet is not None and not '-' in simple_filename:
        index += 1
        worksheet.cell(column=index + 1, row=1, value=f"{simple_filename}")
        worksheet.cell(column=index + 1, row=2, value=f"{len(attributes_used)}")
        worksheet.cell(column=index + 1, row=3, value=f"{len(data)}")
        worksheet.cell(column=index + 1, row=4, value=f"{len(permitted_attributes) - 1}")
        worksheet.cell(column=index + 1, row=5, value=f"{len(permitted_attributes) > 1}")
        if len(permitted_attributes) > 1:
            copyfile(new_filename, f"{base_dir}/{new_filename.rsplit('/')[-1]}")
    return index, worksheet


def apply_cleaning_recursively(root_dir, worksheet=None, index=0, base_dir=None):
    root_dir = os.path.abspath(root_dir)
    for item in os.listdir(root_dir):
        item_full_path = os.path.join(root_dir, item)
        if os.path.isdir(item_full_path):
            index, worksheet = apply_cleaning_recursively(item_full_path, worksheet, index, base_dir)
        elif item_full_path.endswith(".dat") or item_full_path.endswith(".dat"):
            index, worksheet = modify_file(item_full_path, worksheet, index, base_dir)
            print(f"File {item_full_path} with number {index} cleaned")
    return index, worksheet


parser = argparse.ArgumentParser(
    description='cleans a dataset or set or datasets to only contain categorical attributes')
parser.add_argument('file', help="File or directory to be converted")

args = parser.parse_args()

if os.path.isdir(args.file):
    cleaned_datasets_dir = f"{args.file}/CleanedDatasets"
    if os.path.exists(cleaned_datasets_dir):
        rmtree(cleaned_datasets_dir)
    os.mkdir(cleaned_datasets_dir)
    print(f"{args.file} is a directory, looking for .dat files inside")
    wb = Workbook()
    ws = wb.active
    ws["A2"] = "Number of original attributes"
    ws["A3"] = "Number of instances"
    ws["A4"] = "Number of categorical attributes"
    ws["A5"] = "Is going to be used?"
    index = 0
    apply_cleaning_recursively(args.file, ws, index, cleaned_datasets_dir)
    wb.save(filename=f"{args.file}/accumulated.xlsx")
else:
    modify_file(args.file)
