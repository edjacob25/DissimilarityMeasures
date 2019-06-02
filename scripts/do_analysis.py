import argparse
import configparser
import json
import multiprocessing
import os
import subprocess
import time
from datetime import datetime
from shutil import copyfile

import math
import requests
from openpyxl import Workbook
import git
from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, ForeignKey, Boolean
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import relationship, sessionmaker

Base = declarative_base()


class Experiment(Base):
    __tablename__ = 'experiment'

    id = Column(Integer, primary_key=True)
    method = Column(String)
    f_score = Column(Float)
    command_sent = Column(String)
    time_taken = Column(Float)
    k_means_plusplus = Column(Boolean)
    file_name = Column(String)
    comments = Column(String)
    number_of_clusters = Column(Integer)
    start_time = Column(DateTime)

    set_id = Column(Integer, ForeignKey('experiment_set.id'))
    set = relationship("ExperimentSet", back_populates="experiments")


class ExperimentSet(Base):
    __tablename__ = 'experiment_set'

    id = Column(Integer, primary_key=True)
    time = Column(DateTime)
    time_taken = Column(Float)
    number_of_datasets = Column(Integer)
    base_directory = Column(String)
    commit = Column(String)
    experiments = relationship("Experiment", order_by=Experiment.id, back_populates="set")


def get_number_of_clusters(filepath: str):
    with open(filepath) as file:
        for line in file:
            line_upper = line.upper()
            if line_upper.startswith("@ATTRIBUTE") and ("CLASS" in line_upper or "CLUSTER" in line_upper):
                clusters = line.split(" ", 2)[-1]
                return len(clusters.split(","))
            if line_upper.startswith("@DATA"):
                raise Exception("Could not found Class or Cluster attribute")


def remove_attribute(filepath: str, attribute: str):
    attributes = []
    data = []
    permitted_indexes = []
    with open(filepath) as file:
        relation_name = file.readline()
        data_processing = False
        i = 0
        for line in file:
            if not data_processing:
                line_upper = line.upper()
                if line_upper.startswith("@ATTRIBUTE"):
                    if attribute.upper() not in line_upper:
                        attributes.append(line)
                        permitted_indexes.append(i)
                    i += 1
                    continue
                if line_upper.startswith("@DATA"):
                    data_processing = True
            else:
                line_data = [x.lstrip() for x in line.split(',')]
                data.append(line_data)

    with open(filepath, "w") as new_file:
        new_file.write(relation_name)
        for item in attributes:
            new_file.write(item)
        new_file.write("\n")
        new_file.write("@DATA\n")
        for datapoint in data:
            points = [x for i, x in enumerate(datapoint) if i in permitted_indexes]
            result = ",".join(points)
            new_file.write(result)


# TODO: Add option to run other dissimilarity measures
# TODO: Add option to read classpath from the config file
def cluster_dataset(filepath: str, classpath: str = None, no_classpath: bool = False, verbose: bool = False,
                    strategy: str = "A", weight_strategy: str = "N", other_measure: str = None, start_mode: str = "1") \
        -> Experiment:
    clustered_file_path = filepath.replace(".arff", "_clustered.arff")
    command = ["java", "-Xmx8192m"]
    if not no_classpath:
        java_classpath = "/mnt/c/Program Files/Weka-3-9/weka.jar:/home/jacob/wekafiles/packages" \
                         "/SimmilarityMeasuresForCategoricalData/DissimilarityMeasures-0.1.jar"
        if classpath is not None:
            java_classpath = classpath
        command.append("-cp")
        command.append(java_classpath)

    command.append("weka.filters.unsupervised.attribute.AddCluster")
    command.append("-W")

    num_clusters = get_number_of_clusters(filepath)
    if verbose:
        print(f"Number of clusters for {filepath} is {num_clusters}")
    num_procs = multiprocessing.cpu_count()
    distance_function = f"\"weka.core.LearningBasedDissimilarity -R first-last -S {strategy} -w {weight_strategy}\""
    if other_measure is not None:
        distance_function = f"\"{other_measure}\""
    clusterer = f"weka.clusterers.CategoricalKMeans -init {start_mode} -max-candidates 100 -periodic-pruning 10000 " \
        f"-min-density 2.0 -t1 -1.25 -t2 -1.0 -N {num_clusters} -A {distance_function} -I 500 " \
        f"-num-slots {math.floor(num_procs / 3)} -S 10"
    command.append(clusterer)
    command.append("-i")
    command.append(filepath)
    command.append("-o")
    command.append(clustered_file_path)
    command.append("-I")
    command.append("Last")

    start_dt = datetime.now()
    start = time.time()
    result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    end = time.time()

    print(result.stderr.decode("utf-8"))
    if verbose:
        print(result.stdout.decode("utf-8"))
        print(f"Analyzing dataset {filepath} with strategy {strategy} took {end - start}")

    if "Exception" not in result.stderr.decode("utf-8"):
        remove_attribute(clustered_file_path, "Class")
        print(f"Finished clustering dataset {filepath} with strategy {strategy} and weight {weight_strategy}")
    else:
        if os.path.exists(clustered_file_path):
            os.remove(clustered_file_path)

        if start_mode == "1":
            print("There was an error running weka with the k-means++ mode, trying with classic mode")
            return cluster_dataset(filepath, classpath=classpath, no_classpath=no_classpath, verbose=verbose,
                                   strategy=strategy, weight_strategy=weight_strategy, other_measure=other_measure,
                                   start_mode="0")
        else:
            raise Exception(f"There was a error running weka with the file {filepath.rsplit('/')[-1]} and the " +
                            f"following command {' '.join(result.args)}")

    if start_mode == "1":
        return Experiment(method=distance_function.replace("\"", ""), command_sent=" ".join(command),
                          time_taken=end - start,
                          k_means_plusplus=True, file_name=filepath, number_of_clusters=num_clusters,
                          start_time=start_dt, comments="")
    else:
        return Experiment(method=distance_function.replace("\"", ""), command_sent=" ".join(command),
                          time_taken=end - start,
                          k_means_plusplus=False, file_name=filepath, number_of_clusters=num_clusters,
                          start_time=start_dt, comments="")


def copy_files(filepath: str, strategy: str = "", weight_strategy: str = ""):
    path, file = filepath.rsplit("/", 1)
    filename = file.split(".")[0]
    os.mkdir(f"{path}/{filename}_{strategy}_{weight_strategy}")
    new_filepath = f"{path}/{filename}_{strategy}_{weight_strategy}/{file}"
    copyfile(filepath, new_filepath)

    new_clustered_filepath = f"{path}/{filename}_{strategy}_{weight_strategy}/{filename}.clus"
    copyfile(f"{path}/{filename}_clustered.arff", new_clustered_filepath)

    return new_filepath, new_clustered_filepath


def get_f_measure(filepath: str, clustered_filepath: str, exe_path: str = None, verbose: bool = False) -> str:
    command = ["MeasuresComparator.exe", "-c", clustered_filepath, "-r", filepath]
    if exe_path is not None:
        command[0] = exe_path
    start = time.time()
    result = subprocess.run(command, stdout=subprocess.PIPE)
    end = time.time()
    text_result = result.stdout.decode('utf-8')

    if result.returncode != 0:
        print(f"Could not get F-Measure\nError -> {text_result}")
        raise Exception("Could not calculate f-measure")
    else:
        if verbose:
            print(f"Calculating f-measure took {end - start}")
        print(f"Finished getting f-measure for {filepath}, f-measure -> {text_result}")
        return text_result


def send_notification(message: str, title: str):
    config = configparser.ConfigParser()
    config.read("config.ini")
    data = {"body": message, "title": title, "type": "note"}
    headers = {"Content-Type": "application/json", "Access-Token": config["SECRETS"]["Pushbullet_token"]}
    requests.post("https://api.pushbullet.com/v2/pushes", headers=headers, data=json.dumps(data))


def format_seconds(seconds: float) -> str:
    seconds = math.fabs(seconds)
    if seconds > 3600:
        return f"{seconds / 3600} hours"
    elif seconds > 60:
        return f"{seconds / 60} minutes"
    else:
        return f"{seconds} seconds"


def do_analysis(directory: str, verbose: bool, cp: str = None, measure_calculator_path: str = None,
                alternate: bool = False):
    if alternate:
        measures = ["weka.core.Eskin", "weka.core.Gambaryan", "weka.core.Goodall", "weka.core.Lin",
                    "weka.core.OccurenceFrequency", "weka.core.InverseOccurenceFrequency",
                    "weka.core.EuclideanDistance", "weka.core.ManhattanDistance", "weka.core.LinModified",
                    "weka.core.LinModified2", "weka.core.LinModified_Kappa", "weka.core.LinModified_MinusKappa",
                    "weka.core.LinModified_KappaMax"]
        measures = list(zip(measures, [None for _ in range(len(measures))]))
    else:
        strategies = ["A", "B", "C", "D", "E", "N"]
        weights = ["N", "K", "A"]
        measures = list(zip(strategies, weights))

    engine = create_engine('sqlite:///results.db')
    Base.metadata.create_all(engine)
    session_class = sessionmaker(bind=engine)
    session = session_class()
    repo = git.Repo(search_parent_directories=True)
    exp_set = ExperimentSet(time=datetime.now(), base_directory=directory, commit=repo.head.object.hexsha)
    session.add(exp_set)
    session.commit()

    start = time.time()
    root_dir = os.path.abspath(directory)
    i = 0
    for item in os.listdir(root_dir):
        if item.rsplit('.', 1)[-1] == "arff" and "clustered" not in item:
            item_fullpath = os.path.join(root_dir, item)
            try:
                for strategy, weight in measures:
                    if weight is None:
                        exp = cluster_dataset(item_fullpath, verbose=verbose, classpath=cp, other_measure=strategy)
                    else:
                        exp = cluster_dataset(item_fullpath, verbose=verbose, classpath=cp, strategy=strategy,
                                              weight_strategy=weight)
                    new_filepath, new_clustered_filepath = copy_files(item_fullpath, strategy=strategy,
                                                                      weight_strategy=weight)
                    f_measure = get_f_measure(new_filepath, new_clustered_filepath,
                                              exe_path=measure_calculator_path,
                                              verbose=verbose)
                    exp.f_score = f_measure
                    exp_set.experiments.append(exp)

                session.commit()
                i += 1
            except KeyboardInterrupt:
                session.rollback()
                print(f"The analysis of the file {item} was requested to be finished by using Ctrl-C")
                continue
            except Exception as exc:
                session.rollback()
                print(exc)
                print(f"Skipping file {item}")
                continue
            finally:
                print("\n\n")

    end = time.time()

    exp_set.time_taken = end - start
    exp_set.number_of_datasets = i
    session.commit()
    create_report(exp_set.id, base_path=root_dir)

    time_str = format_seconds(end - start)
    send_notification(f"It took {time_str} and processed {i} datasets", "Analysis finished")


def create_report(experiment_set: int, base_path: str = ""):
    wb = Workbook()
    ws = wb.active
    engine = create_engine('sqlite:///results.db')
    session_class = sessionmaker(bind=engine)
    session = session_class()
    headers = []
    row = 1
    last = ""
    column = 2
    for experiment in session.query(Experiment).filter_by(set_id=experiment_set).order_by(Experiment.file_name):
        if last != experiment.file_name:
            column = 2
            row += 1
            last = experiment.file_name
            ws.cell(row=row, column=1, value=experiment.file_name.rsplit('/')[-1])
        if experiment.method not in headers:
            headers.append(experiment.method)
        ws.cell(row=row, column=column, value=experiment.f_score)
        column += 1

    for i, header in enumerate(headers):
        ws.cell(row=1, column=i + 2, value=header)
        ws.cell(row=1, column=i + column + 2, value=header)

    for i in range(2, row + 1):
        base = ord('A') + column - 2
        for j in range(column - 2):
            item = ord('B') + j
            ws.cell(row=i, column=j + column + 2, value=f"=RANK({chr(item)}{i},B{i}:{chr(base)}{i})")

    start = chr(ord('B') + column)
    end = chr(ord('B') + column + column - 3)
    for i in range(column - 2):
        item = chr(ord('B') + i + column)
        ws.cell(row=row + 3, column=i + column + 2, value=f"=AVERAGE({item}2:{item}{row + 1})")
        ws.cell(row=row + 4, column=i + column + 2, value=f"=RANK({item}{row + 3},{start}{row + 3}:{end}{row + 3})")

    save_path = os.path.join(base_path, "results.xlsx")
    print(f"Saving to {save_path}")
    wb.save(save_path)


parser = argparse.ArgumentParser(description='Does the analysis of a directory containing categorical datasets')
parser.add_argument('directory', help="Directory in which the cleaned datasets are")
parser.add_argument('-cp', help="Classpath for the weka invocation, needs to contain the weka.jar file and probably "
                                "the jar of the measure ")
parser.add_argument("-v", "--verbose", help="Show the output of the weka commands", action='store_true')
parser.add_argument("-f", "--measure-calc", help="Path to the f-measure calculator", dest='measure_calculator_path')
parser.add_argument("--alternate-analysis", help="Does the alternate analysis with the already known simmilarity "
                                                 "measures", action='store_true')
parser.add_argument("-s", "--save", help="Path to the f-measure calculator", action='store_true')
# TODO: Actually save the output of the commands


args = parser.parse_args()

# TODO: Read a single file
if not os.path.isdir(args.directory):
    print("The selected path is not a directory")
    exit(1)

do_analysis(args.directory, args.verbose, args.cp, args.measure_calculator_path, args.alternate_analysis)
